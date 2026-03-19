"""엑셀 결과 파일 생성 서비스 (openpyxl 기반)"""
import logging
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session
from app.models import Transaction
from app.parsers.common import is_full_card_number
from app.services.master_service import (
    get_all_projects, get_all_solutions, get_all_account_subjects,
)
from app.config import EXPORT_DIR

logger = logging.getLogger(__name__)

RESULT_COLUMNS = [
    "승인일",
    "승인시간",
    "카드번호",
    "이용자명",
    "가맹점명",
    "업종명",
    "승인금액",
    "프로젝트명",
    "솔루션",
    "계정과목/지출내역",
    "회식비/접대비/회의비 Flex 사전승인 유무",
    "참석자이름(모두기재)",
    "구매내역",
    "기타사항",
]

FLEX_OPTIONS = ["O", "X"]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _get_transactions_for_card(
    db: Session,
    card_identifier: str,
    bank: str,
    year_month: str | None,
    card_number_raw: str | None = None,
) -> list[Transaction]:
    """card_identifier: 전체 카드번호 또는 끝4자리. card_number_raw가 전체번호면 우선 사용."""
    q = db.query(Transaction).filter(Transaction.source_bank == bank)
    if card_number_raw and is_full_card_number(card_number_raw):
        q = q.filter(Transaction.card_number_raw == card_number_raw)
    else:
        q = q.filter(Transaction.card_last4 == card_identifier)
    if year_month:
        q = q.filter(Transaction.use_year_month == year_month)
    return q.order_by(Transaction.approval_datetime).all()


def _build_options_sheet(
    wb: openpyxl.Workbook,
    projects: list[str],
    solutions: list[str],
    accounts: list[str],
) -> None:
    """숨김 시트(meta)에 드롭다운 옵션 목록 저장"""
    ws = wb.create_sheet("meta")
    ws.sheet_state = "hidden"

    ws["A1"] = "프로젝트"
    ws["B1"] = "솔루션"
    ws["C1"] = "계정과목"
    ws["D1"] = "Flex"

    for i, name in enumerate(projects, start=2):
        ws.cell(row=i, column=1, value=name)
    for i, name in enumerate(solutions, start=2):
        ws.cell(row=i, column=2, value=name)
    for i, name in enumerate(accounts, start=2):
        ws.cell(row=i, column=3, value=name)
    for i, val in enumerate(FLEX_OPTIONS, start=2):
        ws.cell(row=i, column=4, value=val)


def _add_data_validations(
    ws,
    data_rows: range,
    projects: list[str],
    solutions: list[str],
    accounts: list[str],
    meta_start_row: int = 2,
) -> None:
    """데이터 유효성 검사(드롭다운) 추가"""
    proj_end = meta_start_row + len(projects) - 1
    sol_end = meta_start_row + len(solutions) - 1
    acc_end = meta_start_row + len(accounts) - 1
    flex_end = meta_start_row + len(FLEX_OPTIONS) - 1

    proj_src = f"meta!$A${meta_start_row}:$A${proj_end}" if projects else None
    sol_src = f"meta!$B${meta_start_row}:$B${sol_end}" if solutions else None
    acc_src = f"meta!$C${meta_start_row}:$C${acc_end}" if accounts else None
    flex_src = f"meta!$D${meta_start_row}:$D${flex_end}"

    # 컬럼 인덱스 (1-based): 프로젝트=8, 솔루션=9, 계정과목=10, Flex=11
    col_configs = [
        (8, proj_src),
        (9, sol_src),
        (10, acc_src),
        (11, flex_src),
    ]

    for col_idx, formula in col_configs:
        if not formula:
            continue
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell_range = f"{col_letter}{data_rows.start}:{col_letter}{data_rows.stop - 1}"
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            error="목록에서 선택하세요.",
            errorTitle="입력 오류",
        )
        ws.add_data_validation(dv)
        dv.sqref = cell_range


def generate_card_excel(
    db: Session,
    card_last4: str,
    bank: str,
    year_month: str | None,
    user_name: str,
    card_number_raw: str | None = None,
) -> Path:
    """카드 단건 엑셀 파일 생성 후 파일 경로 반환"""
    transactions = _get_transactions_for_card(
        db, card_last4, bank, year_month, card_number_raw=card_number_raw
    )

    projects = [p.name for p in get_all_projects(db, active_only=True)]
    solutions = [s.name for s in get_all_solutions(db, active_only=True)]
    accounts = [a.name for a in get_all_account_subjects(db, active_only=True)]

    wb = openpyxl.Workbook()
    _build_options_sheet(wb, projects, solutions, accounts)

    ws = wb.active
    ws.title = "내역"

    # 헤더 작성
    for col_idx, col_name in enumerate(RESULT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # 데이터 작성
    for row_idx, tx in enumerate(transactions, start=2):
        values: list[Any] = [
            tx.approval_date,
            tx.approval_time,
            tx.card_number_raw,
            tx.card_owner_name or "",
            tx.merchant_name or "",
            tx.merchant_category or "",
            tx.approval_amount,
            tx.project_name or "",
            tx.solution_name or "",
            tx.account_subject or "",
            tx.flex_pre_approved or "",
            tx.attendees or "",
            tx.purchase_detail or "",
            tx.remarks or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # 입력 필요 컬럼 배경색
            if col_idx in (8, 9, 10, 11, 12, 13, 14):
                cell.fill = INPUT_FILL

    # 드롭다운 추가
    if len(transactions) > 0:
        data_rows = range(2, len(transactions) + 2)
        _add_data_validations(ws, data_rows, projects, solutions, accounts)

    # 컬럼 너비 자동 조정
    col_widths = [12, 10, 22, 10, 25, 15, 12, 25, 20, 25, 20, 20, 25, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 틀 고정 (헤더)
    ws.freeze_panes = "A2"

    # 파일명: {사용자명} {YYYYMM}({은행코드}).xlsx
    ym = year_month or "000000"
    file_name = f"{user_name} {ym}({bank}).xlsx"
    file_path = EXPORT_DIR / file_name

    wb.save(file_path)
    logger.info(f"엑셀 생성 완료: {file_path}")
    return file_path
